import pandas as pd
from telethon.sync import TelegramClient
from telethon.tl.functions.messages import GetHistoryRequest
import time
import re
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
from datetime import datetime, timedelta

# Настройка логирования для вывода в консоль
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

logging.info('Starting the program')

api_id = '23898969'
api_hash = '38aff42e32aab3cfa6b5128100e300ad'
phone = '+375336404737'
channel_username = 'greekslive_notifications2'
folder_name = 'greekslive_notifications2'

# Создаем папку, если она не существует
if not os.path.exists(folder_name):
    os.makedirs(folder_name)

client = TelegramClient('session_name_greekslive', api_id, api_hash)

logging.info('Connecting to Telegram')

client.connect()

if not client.is_user_authorized():
    client.send_code_request(phone)
    client.sign_in(phone, input('Enter the code: '))

def parse_message(message):
    type_match = re.search(r'([A-Z ]+ \(.+\))', message)
    time_match = re.search(r'Time: (.+)\n', message)
    index_match = re.search(r'Index: (.+)\n', message)
    total_match = re.search(r'Total (Debit|Credit): (.+)\n', message)
    metrics_match = re.search(r'Δ: .+, Γ: .+, ν: .+, Θ: .+', message)
    notional_value_match = re.search(r'Notional Value: (.+)\n', message)
    sfm_combo_match = re.search(r'SFM for Combo: (.+)\n', message)
    sfm_contract_match = re.search(r'SFM per Contract: (.+)\n', message)
    apr_match = re.search(r'APR: (.+)%', message)

    # Получение всех строк между "Index" и "Total"
    order_matches = re.findall(r'(🟢 Buy .+|🔴 Sell .+)', message)
    order_part = '\n'.join(order_matches).strip() if order_matches else None  # Объединяем строки с сохранением переносов

    return {
        'Type': type_match.group(1) if type_match else None,
        'Time': time_match.group(1) if time_match else None,
        'Index': index_match.group(1) if index_match else None,
        'Order': order_part,
        'Total': total_match.group(2) if total_match else None,
        'Metrics': metrics_match.group(0) if metrics_match else None,
        'Notional Value': notional_value_match.group(1) if notional_value_match else None,
        'SFM for Combo': sfm_combo_match.group(1) if sfm_combo_match else None,
        'SFM per Contract': sfm_contract_match.group(1) if sfm_contract_match else None,
        'APR': apr_match.group(1) if apr_match else None,
    }

def adjust_excel_formatting(filename):
    wb = load_workbook(filename)
    ws = wb.active

    # Автоматическая настройка ширины столбцов и высоты строк, а также центрирование текста
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')  # Центрирование и перенос текста
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) if column != 'D' else 60  # Задаем фиксированную ширину для столбца Order (столбец D)
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                cell_value = str(cell.value)
                lines = cell_value.split('\n')
                if len(lines) > max_height:
                    max_height = len(lines)
        ws.row_dimensions[row[0].row].height = max_height * 15  # Примерная высота строки для текста

    wb.save(filename)

def get_excel_filename():
    current_time = datetime.now() + timedelta(hours=3)  # Московское время
    date_str = current_time.strftime('%d-%m-%y')
    return os.path.join(folder_name, f'{date_str}-greekslive_notifications2.xlsx')

# Получение канала по username
channel = client.get_entity(channel_username)

logging.info('Successfully connected to Telegram and retrieved channel information')

# Получаем ID последнего сообщения при старте
history = client(GetHistoryRequest(
    peer=channel,
    limit=1,
    offset_date=None,
    offset_id=0,
    max_id=0,
    min_id=0,
    add_offset=0,
    hash=0
))
last_message_id = history.messages[0].id if history.messages else None

# Получаем текущее время и дату для проверки обновления файла
current_date = (datetime.now() + timedelta(hours=3)).date()
excel_file = get_excel_filename()

while True:
    try:
        history = client(GetHistoryRequest(
            peer=channel,
            limit=100,
            offset_date=None,
            offset_id=0,
            max_id=0,
            min_id=0,
            add_offset=0,
            hash=0
        ))
        if history.messages:
            new_messages = [msg for msg in history.messages if msg.id > last_message_id and msg.message]
            for message in reversed(new_messages):
                if message.message:  # Проверяем, что сообщение является текстовым
                    logging.info('New message received: %s', message.message)
                    data = parse_message(message.message)
                    new_row = pd.DataFrame([data])

                    # Проверяем, если дата изменилась, создаем новый файл
                    new_date = (datetime.now() + timedelta(hours=3)).date()
                    if new_date != current_date:
                        current_date = new_date
                        excel_file = get_excel_filename()
                        logging.info('Creating new file for date: %s', new_date)

                    # Проверка наличия файла и создание нового при отсутствии
                    if not os.path.exists(excel_file):
                        df = pd.DataFrame(columns=['Type', 'Time', 'Index', 'Order', 'Total', 'Metrics', 'Notional Value', 'SFM for Combo', 'SFM per Contract', 'APR'])
                    else:
                        df = pd.read_excel(excel_file)  # Загружаем существующий файл, если он уже есть

                    df = pd.concat([df, new_row], ignore_index=True)
                    df.to_excel(excel_file, index=False)
                    adjust_excel_formatting(excel_file)
                    logging.info('Message parsed and saved to Excel')

                    last_message_id = message.id
    except Exception as e:
        logging.error('An error occurred: %s', str(e))
    
    time.sleep(60)

client.disconnect()
logging.info('Disconnected from Telegram')
