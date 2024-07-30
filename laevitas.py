import pandas as pd
from telethon.sync import TelegramClient
from telethon.tl.functions.messages import GetHistoryRequest
import time
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
from datetime import datetime, timedelta

# Настройка логирования для вывода в консоль
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

logging.info('Starting the program')

api_id = '23898969'
api_hash = '38aff42e32aab3cfa6b5128100e300ad'
phone = '+375336404737'
channel_username = 'laevitas'
folder_name = 'laevitas'

# Создаем папку, если она не существует
if not os.path.exists(folder_name):
    os.makedirs(folder_name)

client = TelegramClient('session_name_laevitas', api_id, api_hash)

logging.info('Connecting to Telegram')

client.connect()

if not client.is_user_authorized():
    client.send_code_request(phone)
    client.sign_in(phone, input('Enter the code: '))

def parse_message(message):
    current_time_msk = (datetime.utcnow() + timedelta(hours=3)).strftime('%Y-%m-%d %H:%M:%S')
    return {
        'Message': message,
        'Time (MSK)': current_time_msk
    }

def adjust_excel_formatting(filename):
    wb = load_workbook(filename)
    ws = wb.active

    # Автоматическая настройка ширины столбцов и высоты строк, а также центрирование текста
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')  # Центрирование и перенос текста
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) if column != 'A' else 60  # Задаем фиксированную ширину для столбца Message (столбец A)
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                cell_value = str(cell.value)
                lines = cell_value.split('\n')
                if len(lines) > max_height:
                    max_height = len(lines)
        ws.row_dimensions[row[0].row].height = max_height * 15  # Примерная высота строки для текста

    wb.save(filename)

def get_excel_filename():
    current_time = datetime.now() + timedelta(hours=3)  # Московское время
    date_str = current_time.strftime('%d-%m-%y')
    return os.path.join(folder_name, f'{date_str}-laevitas.xlsx')

# Получение канала по username
channel = client.get_entity(channel_username)

logging.info('Successfully connected to Telegram and retrieved channel information')

# Получаем ID последнего сообщения при старте
history = client(GetHistoryRequest(
    peer=channel,
    limit=1,
    offset_date=None,
    offset_id=0,
    max_id=0,
    min_id=0,
    add_offset=0,
    hash=0
))
last_message_id = history.messages[0].id if history.messages else None

# Получаем текущее время и дату для проверки обновления файла
current_date = (datetime.now() + timedelta(hours=3)).date()
excel_file = get_excel_filename()

while True:
    try:
        history = client(GetHistoryRequest(
            peer=channel,
            limit=100,
            offset_date=None,
            offset_id=0,
            max_id=0,
            min_id=0,
            add_offset=0,
            hash=0
        ))
        if history.messages:
            new_messages = [msg for msg in history.messages if msg.id > last_message_id and msg.message]
            for message in reversed(new_messages):
                if message.message:  # Проверяем, что сообщение является текстовым
                    logging.info('New message received: %s', message.message)
                    data = parse_message(message.message)
                    new_row = pd.DataFrame([data])

                    # Проверяем, если дата изменилась, создаем новый файл
                    new_date = (datetime.now() + timedelta(hours=3)).date()
                    if new_date != current_date:
                        current_date = new_date
                        excel_file = get_excel_filename()
                        logging.info('Creating new file for date: %s', new_date)

                    # Проверка наличия файла и создание нового при отсутствии
                    if not os.path.exists(excel_file):
                        df = pd.DataFrame(columns=['Message', 'Time (MSK)'])
                    else:
                        df = pd.read_excel(excel_file)  # Загружаем существующий файл, если он уже есть

                    df = pd.concat([df, new_row], ignore_index=True)
                    df.to_excel(excel_file, index=False)
                    adjust_excel_formatting(excel_file)
                    logging.info('Message parsed and saved to Excel')

                    last_message_id = message.id
    except Exception as e:
        logging.error('An error occurred: %s', str(e))
    
    time.sleep(60)

client.disconnect()
logging.info('Disconnected from Telegram')
